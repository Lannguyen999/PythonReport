from openpyxl import Workbook
import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def createWorkbook(sheet_titles)->Workbook:
  wb = Workbook()
  for idx, title in enumerate(sheet_titles):
    if idx == 0:
      ws = wb.active
      ws.title = title
    else:
      ws = wb.create_sheet(title)
  return wb

def createTitle(
  ws, 
  row:int, 
  start_col:int, 
  end_col:int, 
  title:str, 
  align:str = "center"
):
  text = ws.cell(row = row, column = start_col, value = title)
  text.font = Font(name = "Times New Roman",size= 14, bold=True)
  text.alignment = Alignment(horizontal=align)
  ws.merge_cells(start_row= row, start_column= start_col, end_row= row, end_column= end_col)
          
def setCellText(
  ws, 
  row:int, 
  col:int, 
  text:str, 
  bold:bool = False, 
  italic:bool = False, 
  align:str="left", 
  borderStyle:str|None = None,
  fontColor:str = "000000",
  cellColor:str = "FFFFFF"
):
  tx = ws.cell(row = row, column = col, value = text)
  # Available Border Style: thin, double, thick
  tx.font = Font(name="Times New Roman", bold=bold, italic=italic, color=fontColor)
  tx.alignment = Alignment(horizontal=align)
  if borderStyle is not None:
    border = Side(border_style=borderStyle, color="000000")
    tx.border = Border(top=border, left=border, bottom=border, right=border)
  tx.fill = PatternFill("solid", fgColor=cellColor)
  
def setHeader(
  ws,
  start_row,
  end_row,
  start_col,
  end_col,
  text,
  bold:bool = True,
  italic:bool = False,
  align:str = "center",
  borderStyle:str = "thin",
  fontColor:str = "000000",
  cellColor:str = "FFFFFF"
):
  border = Side(border_style=borderStyle, color="000000")
  for r in range(start_row, end_row + 1):
    for c in range(start_col, end_col + 1):
        tx = ws.cell(row = r, column = c, value=None)
        tx.border = Border(top=border, left=border, bottom=border, right=border)
        tx.fill = PatternFill("solid", fgColor=cellColor)
  tx = ws.cell(row = start_row, column = start_col, value=text)
  tx.font = Font(name="Times New Roman", bold=bold, italic=italic, color=fontColor)
  tx.alignment = Alignment(horizontal=align, vertical="center")
  ws.merge_cells(start_row= start_row, start_column= start_col, end_row= end_row, end_column= end_col)
  
def setFormula(
  ws,
  start_row,
  end_row,
  start_col,
  end_col,
  formula,
  bold:bool = True,
  align:str = "right",
  number_format:str = "###,###,###",
  borderStyle:str = "thin",
  fontColor:str = "000000",
  cellColor:str = "FFFFFF"
):
  border = Side(border_style=borderStyle, color="000000")
  for r in range(start_row, end_row + 1):
    for c in range(start_col, end_col + 1):
        tx = ws.cell(row = r, column = c, value=None)
        tx.border = Border(top=border, left=border, bottom=border, right=border)
        tx.fill = PatternFill("solid", fgColor=cellColor)
  tx = ws.cell(row = start_row, column = start_col, value=f"={formula}")
  tx.font = Font(name="Times New Roman", bold=bold, color=fontColor)
  tx.alignment = Alignment(horizontal=align, vertical="center")
  tx.number_format = number_format
  ws.merge_cells(start_row= start_row, start_column= start_col, end_row= end_row, end_column= end_col)
  
def setData(df,ws, start_row, start_col, header:bool=False, index:bool=False):
  rows = dataframe_to_rows(df, header=header, index=index)
  for r_idx, row in enumerate(rows, start_row):
    for c_idx, value in enumerate(row, start_col + 1):
      v = ws.cell(row= r_idx, column = c_idx, value = value)
      v.font = Font(name="Times New Roman")
      border = Side(border_style="thin", color="000000")
      v.border = Border(top=border, left=border, bottom=border, right=border)
      if type(value) is int:
        v.number_format = "###,###,###"
      elif type(value) is datetime.date:
        v.number_format = "dd/mm/yyyy"
        
def setSheetColumnsWidth(ws, widths):
  for i, column_width in enumerate(widths,1):  # ,1 to start at 1
    ws.column_dimensions[get_column_letter(i)].width = column_width